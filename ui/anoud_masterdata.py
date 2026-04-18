"""
Lookup helpers for Anoudapps/QIC tariff codes.

We get Bayanaty IDs from `bayanaty/vehicleDetails`, but the tariff API expects different
`makeCode` / `modelCode` values. This module loads the mapping from the provided Excel
master data file and caches it in-process.
"""

from __future__ import annotations

import logging
import os
from typing import Dict, Optional, Tuple

logger = logging.getLogger(__name__)

_CACHE: Optional[Tuple[Dict[str, str], Dict[str, str]]] = None
_CACHE_PATH: Optional[str] = None


def _find_col(headers: Dict[int, str], *needles: str) -> Optional[int]:
    """
    Find a column index whose header contains all needle substrings (case-insensitive).
    """
    n = [x.lower().strip() for x in needles if x]
    for idx, h in headers.items():
        hh = (h or "").lower()
        if all(x in hh for x in n):
            return idx
    return None


def load_make_model_mappings(xlsx_path: str) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Returns (bayanaty_make_id -> tariff_make_code, bayanaty_model_id -> tariff_model_code).
    Best-effort header detection to survive column name changes.
    """
    global _CACHE, _CACHE_PATH
    if _CACHE is not None and _CACHE_PATH == xlsx_path:
        return _CACHE

    make_map: Dict[str, str] = {}
    model_map: Dict[str, str] = {}

    if not xlsx_path or not os.path.exists(xlsx_path):
        _CACHE = (make_map, model_map)
        _CACHE_PATH = xlsx_path
        return _CACHE

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        logger.warning("openpyxl not available; cannot load Anoudapps masterdata | %s", e)
        _CACHE = (make_map, model_map)
        _CACHE_PATH = xlsx_path
        return _CACHE

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        # Read header row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            _CACHE = (make_map, model_map)
            _CACHE_PATH = xlsx_path
            return _CACHE

        headers: Dict[int, str] = {i: (str(v).strip() if v is not None else "") for i, v in enumerate(header_row)}

        # Try to locate columns (many variants exist in customer files)
        col_b_make = (
            _find_col(headers, "bayanaty", "make", "id")
            or _find_col(headers, "make", "bayanaty")
            or _find_col(headers, "bayanaty", "make")
        )
        col_b_model = (
            _find_col(headers, "bayanaty", "model", "id")
            or _find_col(headers, "model", "bayanaty")
            or _find_col(headers, "bayanaty", "model")
        )
        col_t_make = (
            _find_col(headers, "makecode")
            or _find_col(headers, "make", "code")
            or _find_col(headers, "tariff", "make")
            or _find_col(headers, "qic", "make")
        )
        col_t_model = (
            _find_col(headers, "modelcode")
            or _find_col(headers, "model", "code")
            or _find_col(headers, "tariff", "model")
            or _find_col(headers, "qic", "model")
        )

        if col_b_make is None or col_b_model is None or col_t_make is None or col_t_model is None:
            logger.warning(
                "Could not detect columns in masterdata | path=%s | headers=%s",
                xlsx_path,
                list(headers.values())[:50],
            )
            _CACHE = (make_map, model_map)
            _CACHE_PATH = xlsx_path
            return _CACHE

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                b_make = row[col_b_make]
                b_model = row[col_b_model]
                t_make = row[col_t_make]
                t_model = row[col_t_model]
            except Exception:
                continue

            b_make_s = str(b_make).strip() if b_make is not None else ""
            b_model_s = str(b_model).strip() if b_model is not None else ""
            t_make_s = str(t_make).strip() if t_make is not None else ""
            t_model_s = str(t_model).strip() if t_model is not None else ""

            if b_make_s and t_make_s:
                make_map.setdefault(b_make_s, t_make_s)
            if b_model_s and t_model_s:
                model_map.setdefault(b_model_s, t_model_s)

        _CACHE = (make_map, model_map)
        _CACHE_PATH = xlsx_path
        logger.info(
            "Loaded Anoudapps masterdata mapping | makes=%d models=%d | path=%s",
            len(make_map),
            len(model_map),
            xlsx_path,
        )
        return _CACHE
    except Exception as e:
        logger.exception("Failed loading Anoudapps masterdata | path=%s | %s", xlsx_path, e)
        _CACHE = (make_map, model_map)
        _CACHE_PATH = xlsx_path
        return _CACHE

