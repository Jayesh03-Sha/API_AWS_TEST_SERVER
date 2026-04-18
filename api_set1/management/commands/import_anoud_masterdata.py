from __future__ import annotations

import os
from typing import Dict, Optional

from django.core.management.base import BaseCommand


class Command(BaseCommand):
    help = "Import Bayanaty→tariff make/model mappings from Excel into DB."

    def add_arguments(self, parser):
        parser.add_argument(
            "--xlsx",
            default=os.environ.get(
                "ANOUDAPPS_MAKE_MODEL_XLSX",
                "/Users/sureshkamal/Desktop/Promise_insurance_service/QIC/Make_Model_list_with_Bayanaty_MasterData.xlsx",
            ),
            help="Path to Make_Model_list_with_Bayanaty_MasterData.xlsx",
        )

    def handle(self, *args, **options):
        from api_set1.models import AnoudMakeModelMapping

        xlsx_path = options["xlsx"]
        if not xlsx_path or not os.path.exists(xlsx_path):
            self.stderr.write(self.style.ERROR(f"Excel not found: {xlsx_path}"))
            return

        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as e:
            self.stderr.write(self.style.ERROR(f"openpyxl not installed: {e}"))
            return

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            self.stderr.write(self.style.ERROR("Empty sheet"))
            return

        headers: Dict[int, str] = {i: (str(v).strip() if v is not None else "") for i, v in enumerate(header_row)}

        def find_col(*needles: str) -> Optional[int]:
            ns = [n.lower().strip() for n in needles if n]
            for idx, h in headers.items():
                hh = (h or "").lower()
                if all(n in hh for n in ns):
                    return idx
            return None

        col_b_make = find_col("bayanaty", "make", "id") or find_col("bayanaty", "make")
        col_b_model = find_col("bayanaty", "model", "id") or find_col("bayanaty", "model")
        col_t_make = find_col("makecode") or find_col("make", "code")
        col_t_model = find_col("modelcode") or find_col("model", "code")
        col_make_name = find_col("make", "name") or find_col("make")
        col_model_name = find_col("model", "name") or find_col("model")

        if col_b_make is None or col_b_model is None or col_t_make is None or col_t_model is None:
            self.stderr.write(self.style.ERROR(f"Could not detect required columns. Headers: {list(headers.values())[:50]}"))
            return

        created = 0
        updated = 0
        total = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            total += 1
            b_make = row[col_b_make] if col_b_make < len(row) else None
            b_model = row[col_b_model] if col_b_model < len(row) else None
            t_make = row[col_t_make] if col_t_make < len(row) else None
            t_model = row[col_t_model] if col_t_model < len(row) else None

            b_make_s = str(b_make).strip() if b_make is not None else ""
            b_model_s = str(b_model).strip() if b_model is not None else ""
            t_make_s = str(t_make).strip() if t_make is not None else ""
            t_model_s = str(t_model).strip() if t_model is not None else ""

            if not (b_make_s and b_model_s and t_make_s and t_model_s):
                continue

            make_name = None
            model_name = None
            if col_make_name is not None and col_make_name < len(row):
                mn = row[col_make_name]
                make_name = str(mn).strip() if mn is not None else None
            if col_model_name is not None and col_model_name < len(row):
                mdn = row[col_model_name]
                model_name = str(mdn).strip() if mdn is not None else None

            obj, was_created = AnoudMakeModelMapping.objects.update_or_create(
                bayanaty_make_id=b_make_s,
                bayanaty_model_id=b_model_s,
                defaults={
                    "tariff_make_code": t_make_s,
                    "tariff_model_code": t_model_s,
                    "make_name": make_name,
                    "model_name": model_name,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Imported Anoud masterdata | total_rows={total} created={created} updated={updated} | xlsx={xlsx_path}"
            )
        )

