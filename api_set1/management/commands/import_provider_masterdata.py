from __future__ import annotations

import csv
import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import connection, transaction


INDEX_TOKENS = (
    "code",
    "make_code",
    "model_code",
    "bayanaty_make_code",
    "bayanaty_model_code",
)


def _norm_ident(s: str, *, fallback: str) -> str:
    """
    Convert a human header into a SQLite-safe identifier.
    Rules:
    - lowercase
    - spaces => underscore
    - remove special chars
    - collapse underscores
    """
    raw = (s or "").strip().lower()
    raw = raw.replace("\u00a0", " ")  # non-breaking space
    raw = re.sub(r"\s+", "_", raw)
    raw = re.sub(r"[^a-z0-9_]", "", raw)
    raw = re.sub(r"_+", "_", raw).strip("_")
    return raw or fallback


def _dedupe(names: Sequence[str]) -> List[str]:
    out: List[str] = []
    seen: Dict[str, int] = {}
    for n in names:
        base = n
        if base not in seen:
            seen[base] = 1
            out.append(base)
            continue
        seen[base] += 1
        out.append(f"{base}_{seen[base]}")
    return out


def _sqlite_quote_ident(ident: str) -> str:
    # SQLite uses double-quotes for identifiers. Escape inner quotes by doubling them.
    return '"' + ident.replace('"', '""') + '"'


def _is_empty_cell(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _coerce_cell(v: Any) -> Any:
    """
    Coerce Excel-native types into SQLite-friendly types while preserving meaning.
    - dates/datetimes => ISO text
    - bool => 0/1 integer
    - leave numbers as-is
    - leave text as-is (trim only whitespace-only -> NULL)
    """
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s != "" else None
    if isinstance(v, bool):
        return 1 if v else 0
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def _try_parse_csv_scalar(s: str) -> Any:
    """
    Very small "best effort" parser for CSV cells:
    - empty => NULL
    - true/false/yes/no => 1/0
    - int => int
    - float => float
    - ISO date/datetime => ISO string (kept as TEXT)
    - otherwise => original string
    """
    raw = (s or "").strip()
    if raw == "":
        return None
    low = raw.lower()
    if low in {"true", "yes", "y"}:
        return 1
    if low in {"false", "no", "n"}:
        return 0
    # int
    if re.fullmatch(r"[+-]?\d+", raw):
        try:
            return int(raw)
        except Exception:
            pass
    # float
    if re.fullmatch(r"[+-]?\d+\.\d+", raw):
        try:
            return float(raw)
        except Exception:
            pass
    # ISO-ish date/datetime
    try:
        dt = datetime.fromisoformat(raw)
        return dt.isoformat()
    except Exception:
        pass
    try:
        d = date.fromisoformat(raw)
        return d.isoformat()
    except Exception:
        pass
    return raw


def _infer_sqlite_type(values: Iterable[Any]) -> str:
    """
    Infer a reasonable SQLite column affinity.
    SQLite is flexible, but this keeps numbers as numbers when possible.
    """
    has_text = False
    has_real = False
    has_int = False
    for v in values:
        if v is None:
            continue
        if isinstance(v, (int, bool)):
            has_int = True
            continue
        if isinstance(v, float):
            has_real = True
            continue
        # date/datetime already coerced to ISO string => TEXT
        if isinstance(v, str):
            has_text = True
            continue
        # unknown type => TEXT safest
        has_text = True
    if has_text:
        return "TEXT"
    if has_real:
        return "REAL"
    if has_int:
        return "INTEGER"
    return "TEXT"


@dataclass
class TableReport:
    provider: str
    source_file: str
    sheet: str
    table: str
    columns: List[str] = field(default_factory=list)
    row_count: int = 0


class Command(BaseCommand):
    help = "Import provider masterdata (Excel/CSV) into dynamic SQLite tables."

    def add_arguments(self, parser):
        parser.add_argument(
            "--base-dir",
            default=str(getattr(settings, "BASE_DIR", "")) or os.getcwd(),
            help="Project base directory (defaults to Django BASE_DIR).",
        )
        parser.add_argument(
            "--providers-root",
            default=None,
            help=(
                "Root folder containing providers, e.g. "
                "<base>/data/providers. If omitted, uses <base>/data/providers."
            ),
        )
        parser.add_argument(
            "--providers",
            default="DIC,NIA,Annoudapps",
            help="Comma-separated provider folder names under providers root.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Scan and report what would be created, without changing the database.",
        )

    def handle(self, *args, **options):
        base_dir = options["base_dir"]
        providers_root = options["providers_root"] or os.path.join(base_dir, "data", "providers")
        provider_names = [p.strip() for p in (options["providers"] or "").split(",") if p.strip()]
        dry_run = bool(options["dry_run"])

        errors: List[str] = []
        scanned_files: List[str] = []
        detected_sheets: List[Tuple[str, str, List[str]]] = []
        table_reports: List[TableReport] = []

        # Validate openpyxl lazily only if we see Excel files.
        openpyxl_load_workbook = None

        for provider_folder in provider_names:
            provider_key = _norm_ident(provider_folder, fallback=provider_folder.lower())
            masterdata_dir = os.path.join(providers_root, provider_folder, "masterdata")
            if not os.path.isdir(masterdata_dir):
                errors.append(f"Missing folder: {masterdata_dir}")
                continue

            for root, _dirs, files in os.walk(masterdata_dir):
                for fn in sorted(files):
                    if fn.startswith("~$"):
                        continue  # Excel temp file
                    path = os.path.join(root, fn)
                    ext = os.path.splitext(fn)[1].lower()
                    if ext not in {".xlsx", ".xlsm", ".csv"}:
                        continue
                    scanned_files.append(path)

                    try:
                        if ext == ".csv":
                            tables = self._import_csv(
                                provider_key=provider_key,
                                file_path=path,
                                dry_run=dry_run,
                                errors=errors,
                            )
                            for rep in tables:
                                table_reports.append(rep)
                            detected_sheets.append((path, "(csv)", ["(single)"]))
                            continue

                        # Excel
                        if openpyxl_load_workbook is None:
                            try:
                                from openpyxl import load_workbook  # type: ignore
                            except Exception as e:
                                raise RuntimeError(f"openpyxl not installed: {e}")
                            openpyxl_load_workbook = load_workbook

                        wb = openpyxl_load_workbook(path, read_only=True, data_only=True)
                        sheet_names = list(wb.sheetnames or [])
                        detected_sheets.append((path, "(xlsx)", sheet_names))

                        if len(sheet_names) <= 1:
                            # Rule: if workbook has only one sheet, table name uses filename.
                            table_name = self._make_table_name(
                                provider_key=provider_key,
                                from_filename=os.path.splitext(os.path.basename(path))[0],
                            )
                            sheet_name = sheet_names[0] if sheet_names else "Sheet1"
                            ws = wb[sheet_name]
                            rep = self._import_worksheet(
                                provider_key=provider_key,
                                file_path=path,
                                sheet_name=sheet_name,
                                table_name=table_name,
                                iter_rows=ws.iter_rows(values_only=True),
                                dry_run=dry_run,
                                errors=errors,
                            )
                            if rep:
                                table_reports.append(rep)
                        else:
                            # Rule: if multiple sheets, table name uses sheet name (source of truth).
                            for sheet_name in sheet_names:
                                ws = wb[sheet_name]
                                table_name = self._make_table_name(provider_key=provider_key, from_sheet=sheet_name)
                                rep = self._import_worksheet(
                                    provider_key=provider_key,
                                    file_path=path,
                                    sheet_name=sheet_name,
                                    table_name=table_name,
                                    iter_rows=ws.iter_rows(values_only=True),
                                    dry_run=dry_run,
                                    errors=errors,
                                )
                                if rep:
                                    table_reports.append(rep)
                    except Exception as e:
                        errors.append(f"{path}: {e}")

        # Output report (requested A-F)
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("A) Files scanned"))
        if scanned_files:
            for p in scanned_files:
                self.stdout.write(f"- {p}")
        else:
            self.stdout.write("(none)")

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("B) Sheet names detected"))
        if detected_sheets:
            for file_path, ftype, sheets in detected_sheets:
                self.stdout.write(f"- {file_path} {ftype}")
                for s in sheets:
                    self.stdout.write(f"  - {s}")
        else:
            self.stdout.write("(none)")

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("C) Tables created / rebuilt"))
        if table_reports:
            for rep in table_reports:
                self.stdout.write(
                    f"- {rep.table}  (provider={rep.provider} file={os.path.basename(rep.source_file)} sheet={rep.sheet})"
                )
        else:
            self.stdout.write("(none)")

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("D) Rows imported per table"))
        if table_reports:
            for rep in table_reports:
                self.stdout.write(f"- {rep.table}: {rep.row_count} rows")
        else:
            self.stdout.write("(none)")

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("E) Errors found"))
        if errors:
            for e in errors:
                self.stdout.write(self.style.WARNING(f"- {e}"))
        else:
            self.stdout.write("(none)")

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("F) How to rerun import"))
        if dry_run:
            self.stdout.write("You ran in --dry-run mode. To actually import, rerun without --dry-run:")
        self.stdout.write("python manage.py import_provider_masterdata")
        self.stdout.write("")

    def _make_table_name(self, *, provider_key: str, from_filename: Optional[str] = None, from_sheet: Optional[str] = None) -> str:
        if from_sheet is not None:
            return _norm_ident(f"{provider_key}_{from_sheet}", fallback=f"{provider_key}_sheet")
        if from_filename is not None:
            return _norm_ident(f"{provider_key}_{from_filename}", fallback=f"{provider_key}_file")
        return _norm_ident(provider_key, fallback="provider")

    def _import_csv(
        self,
        *,
        provider_key: str,
        file_path: str,
        dry_run: bool,
        errors: List[str],
    ) -> List[TableReport]:
        table_name = self._make_table_name(provider_key=provider_key, from_filename=os.path.splitext(os.path.basename(file_path))[0])

        try:
            with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if not header:
                    errors.append(f"{file_path}: empty CSV")
                    return []

                raw_headers = [str(h).strip() if h is not None else "" for h in header]
                norm_headers = _dedupe([_norm_ident(h, fallback=f"col_{i+1}") for i, h in enumerate(raw_headers)])

                rows: List[List[Any]] = []
                for r in reader:
                    if not r:
                        continue
                    parsed = [_try_parse_csv_scalar(v) for v in r]
                    # pad / trim to header length
                    if len(parsed) < len(norm_headers):
                        parsed = parsed + [None] * (len(norm_headers) - len(parsed))
                    else:
                        parsed = parsed[: len(norm_headers)]
                    if all(_is_empty_cell(v) for v in parsed):
                        continue
                    rows.append(parsed)

                rep = self._rebuild_table_and_insert(
                    provider_key=provider_key,
                    source_file=file_path,
                    sheet_name="(csv)",
                    table_name=table_name,
                    columns=norm_headers,
                    rows=rows,
                    dry_run=dry_run,
                )
                return [rep] if rep else []
        except Exception as e:
            errors.append(f"{file_path}: {e}")
            return []

    def _import_worksheet(
        self,
        *,
        provider_key: str,
        file_path: str,
        sheet_name: str,
        table_name: str,
        iter_rows: Iterable[Sequence[Any]],
        dry_run: bool,
        errors: List[str],
    ) -> Optional[TableReport]:
        it = iter(iter_rows)
        header_row = next(it, None)
        if not header_row:
            errors.append(f"{file_path} [{sheet_name}]: empty sheet")
            return None

        raw_headers = [str(v).strip() if v is not None else "" for v in header_row]
        norm_headers = _dedupe([_norm_ident(h, fallback=f"col_{i+1}") for i, h in enumerate(raw_headers)])

        rows: List[List[Any]] = []
        for r in it:
            if not r:
                continue
            coerced = [_coerce_cell(v) for v in r]
            if len(coerced) < len(norm_headers):
                coerced = coerced + [None] * (len(norm_headers) - len(coerced))
            else:
                coerced = coerced[: len(norm_headers)]
            if all(_is_empty_cell(v) for v in coerced):
                continue
            rows.append(coerced)

        return self._rebuild_table_and_insert(
            provider_key=provider_key,
            source_file=file_path,
            sheet_name=sheet_name,
            table_name=table_name,
            columns=norm_headers,
            rows=rows,
            dry_run=dry_run,
        )

    def _rebuild_table_and_insert(
        self,
        *,
        provider_key: str,
        source_file: str,
        sheet_name: str,
        table_name: str,
        columns: List[str],
        rows: List[List[Any]],
        dry_run: bool,
    ) -> Optional[TableReport]:
        if not columns:
            return None

        # Infer SQLite types per column (based on coerced values).
        col_values: List[List[Any]] = [[] for _ in columns]
        for r in rows:
            for i, v in enumerate(r):
                col_values[i].append(v)
        col_types = [_infer_sqlite_type(vals) for vals in col_values]

        q_table = _sqlite_quote_ident(table_name)
        col_defs = ", ".join(
            f"{_sqlite_quote_ident(col)} {col_types[i]} NULL"
            for i, col in enumerate(columns)
        )
        create_sql = f"CREATE TABLE {q_table} ({col_defs});"

        if dry_run:
            return TableReport(
                provider=provider_key,
                source_file=source_file,
                sheet=sheet_name,
                table=table_name,
                columns=columns,
                row_count=len(rows),
            )

        with transaction.atomic():
            with connection.cursor() as cur:
                cur.execute(f"DROP TABLE IF EXISTS {q_table};")
                cur.execute(create_sql)

                if rows:
                    placeholders = ", ".join(["?"] * len(columns))
                    insert_sql = f"INSERT INTO {q_table} ({', '.join(_sqlite_quote_ident(c) for c in columns)}) VALUES ({placeholders});"
                    cur.executemany(insert_sql, rows)

                # Index rules (create indexes for columns containing any token)
                for col in columns:
                    col_l = col.lower()
                    if any(tok in col_l for tok in INDEX_TOKENS):
                        idx_name = _norm_ident(f"idx_{table_name}_{col}", fallback=f"idx_{table_name}")
                        q_idx = _sqlite_quote_ident(idx_name)
                        q_col = _sqlite_quote_ident(col)
                        cur.execute(f"DROP INDEX IF EXISTS {q_idx};")
                        cur.execute(f"CREATE INDEX {q_idx} ON {q_table} ({q_col});")

        return TableReport(
            provider=provider_key,
            source_file=source_file,
            sheet=sheet_name,
            table=table_name,
            columns=columns,
            row_count=len(rows),
        )

